import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import pandas as pd
from openpyxl import load_workbook

# Configuración de la ventana principal
tk_root = tk.Tk()
tk_root.title("Transformador de Números")
tk_root.geometry("1000x600")
tk_root.resizable(False, False)

archivo_excel = None  # Ruta del archivo Excel

# Función para importar un archivo Excel
def importar_excel():
    global archivo_excel
    archivo_excel = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    if archivo_excel:
        try:
            xls = pd.ExcelFile(archivo_excel)
            messagebox.showinfo("Importación", f"Archivo Excel importado correctamente. Hojas disponibles: {xls.sheet_names}")
            boton_registrar.config(state=tk.NORMAL)  # Habilitar el botón de registrar
            actualizar_tabla()  # Mostrar los datos actuales del Excel
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
            archivo_excel = None

# Función para transformar un número
def transformar_numero(numero):
    registros = []
    digitos = [int(d) for d in str(numero)]
    digitos = [(d + 5) % 10 for d in digitos]
    registros.append(digitos[:])
    nuevo_orden = [digitos[0], digitos[2], digitos[3], digitos[1]]
    nuevo_orden[0] = (nuevo_orden[0] + 5) % 10
    registros.append(nuevo_orden[:])
    nuevo_orden[0] = (nuevo_orden[0] + 5) % 10
    nuevo_orden[2], nuevo_orden[3] = nuevo_orden[3], nuevo_orden[2]
    registros.append(nuevo_orden[:])
    nuevo_orden = [(d + 5) % 10 for d in registros[1]]
    registros.append(nuevo_orden[:])
    nuevo_orden = [(d + 5) % 10 for d in registros[2]]
    registros.append(nuevo_orden[:])
    return registros

# Función para registrar y guardar en Excel
def registrar_y_guardar():
    if not archivo_excel:
        messagebox.showerror("Error", "Primero debes importar un archivo Excel.")
        return

    numero = entrada_numero.get()
    loteria = entrada_loteria.get()

    if not numero.isdigit() or len(numero) != 4:
        messagebox.showerror("Error", "Debes ingresar un número de 4 cifras.")
        return

    registros = transformar_numero(numero)

    try:
        wb = load_workbook(archivo_excel)
        if "Resultados" not in wb.sheetnames:
            ws = wb.create_sheet("Resultados")
        else:
            ws = wb["Resultados"]

        col = ws.max_column + 2  # Encontrar la siguiente columna disponible

        ws.cell(row=2, column=col, value="Lotería")
        ws.cell(row=2, column=col+1, value=loteria)
        ws.cell(row=4, column=col, value="Número")
        ws.cell(row=4, column=col+1, value=numero)

        for i, registro in enumerate(registros, start=6):
            ws.cell(row=i, column=col, value="Datos")
            ws.cell(row=i, column=col+1, value="".join(map(str, registro)))

        wb.save(archivo_excel)

        # Mostrar mensaje en verde
        mensaje_estado.config(text="✅ Registro guardado en Excel", fg="green")

        # Mostrar en la interfaz
        resultado_texto.delete("1.0", tk.END)  # Limpiar historial antes de mostrar nuevo
        resultado_texto.insert(tk.END, f"Lotería: {loteria}, Número: {numero}\n")
        for i, registro in enumerate(registros, 1):
            resultado_texto.insert(tk.END, f"Paso {i}: {''.join(map(str, registro))}\n")
        resultado_texto.insert(tk.END, "--------------------\n")

        # Actualizar la tabla con los nuevos datos
        actualizar_tabla()

        # Limpiar los campos de entrada
        entrada_numero.delete(0, tk.END)
        entrada_loteria.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar en Excel: {e}")

# Función para actualizar la tabla con datos del Excel
def actualizar_tabla():
    if not archivo_excel:
        return

    try:
        df = pd.read_excel(archivo_excel, sheet_name="Resultados", header=None)

        for row in tree.get_children():
            tree.delete(row)  # Limpiar tabla

        for i in range(0, len(df.columns), 2):
            if i + 1 < len(df.columns):
                loteria = df.iloc[1, i+1] if not pd.isna(df.iloc[1, i+1]) else "N/A"
                numero = df.iloc[3, i+1] if not pd.isna(df.iloc[3, i+1]) else "N/A"
                tree.insert("", tk.END, values=(loteria, numero))

    except Exception as e:
        print(f"Error al actualizar la tabla: {e}")

# Función para cerrar el programa
def cerrar_programa():
    tk_root.destroy()

# Diseño de la interfaz
frame_principal = tk.Frame(tk_root)
frame_principal.pack(padx=20, pady=20)

# Botón para importar Excel (Arriba del todo)
boton_importar = tk.Button(frame_principal, text="Importar Excel", command=importar_excel, font=("Arial", 12), bg="#3498db", fg="white")
boton_importar.pack(pady=10)

# Sección de ingreso de datos
tk.Label(frame_principal, text="Ingrese la lotería:").pack(pady=5)
entrada_loteria = tk.Entry(frame_principal, font=("Arial", 14))
entrada_loteria.pack()

tk.Label(frame_principal, text="Ingrese un número de 4 cifras:").pack(pady=5)
entrada_numero = tk.Entry(frame_principal, font=("Arial", 14), justify="center")
entrada_numero.pack()

# Botón único para registrar y guardar en Excel
boton_registrar = tk.Button(frame_principal, text="Registrar y Transformar", command=registrar_y_guardar, font=("Arial", 12), bg="#2ecc71", fg="white", state=tk.DISABLED)
boton_registrar.pack(pady=10)

# Mensaje de estado
mensaje_estado = tk.Label(frame_principal, text="", font=("Arial", 12))
mensaje_estado.pack(pady=5)

# Sección del historial de números procesados
tk.Label(frame_principal, text="Historial de números procesados").pack()
resultado_texto = tk.Text(frame_principal, height=8, width=50)
resultado_texto.pack()

# Tabla para mostrar registros en Excel
tk.Label(frame_principal, text="Registros en Excel").pack(pady=5)
tree = ttk.Treeview(frame_principal, columns=("Lotería", "Número"), show="headings")
tree.heading("Lotería", text="Lotería")
tree.heading("Número", text="Número")
tree.pack(pady=10)

# Configurar el cierre del programa
tk_root.protocol("WM_DELETE_WINDOW", cerrar_programa)

# Iniciar la aplicación
tk_root.mainloop()
